import pandas as pd
import openpyxl
from openpyxl.styles import Font
import os
import sys
import io

# 设置标准输出编码为UTF-8，解决Windows中文显示问题
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')


class ExcelCompareTool:
    def __init__(self):
        # 可配置参数
        self.file1_path = "table1.xlsx"  # 第一个Excel文件路径
        self.file2_path = "table2.xlsx"  # 第二个Excel文件路径
        self.unique_column = "A"  # 用作唯一标识的列（可修改）
        self.compare_columns = ["B", "E", "F"]  # 需要比较的列

        # 输出文件路径
        self.output_only_in_table1 = "only_in_table1.xlsx"
        self.output_only_in_table2 = "only_in_table2.xlsx"

    def create_sample_files(self):
        """创建两个样例Excel文件用于测试"""
        print("正在创建样例Excel文件...")

        # 创建表格1的样例数据
        data1 = {
            'A': ['ID001', 'ID002', 'ID003', 'ID004', 'ID005'],
            'B': ['苹果', '香蕉', '橙子', '葡萄', '西瓜'],
            'C': ['红色', '黄色', '橙色', '紫色', '绿色'],
            'D': ['水果A', '水果B', '水果C', '水果D', '水果E'],
            'E': ['10元', '5元', '8元', '15元', '20元'],
            'F': ['北京', '上海', '广州', '深圳', '杭州']
        }

        df1 = pd.DataFrame(data1)
        df1.to_excel(self.file1_path, index=False)

        # 添加删除线格式到表1的某些行（用于测试）
        self._add_strikethrough_format_table1()

        # 创建表格2的样例数据（包含一些差异和删除线行）
        data2 = {
            'A': ['ID001', 'ID002', 'ID003', 'ID004', 'ID006', 'ID007', 'ID008'],
            'B': ['苹果', '香蕉', '橙子', '芒果', '桃子', '柠檬', '草莓'],  # ID004的B列不同
            'C': ['红色', '黄色', '橙色', '黄色', '粉色', '黄色', '绿色'],
            'D': ['水果A', '水果B', '水果C', '水果F', '水果G', '水果H', '水果I'],
            'E': ['12元', '5元', '9元', '18元', '25元', '6元', '30元'],  # ID001, ID003, ID004的E列不同
            'F': ['北京', '上海', '深圳', '成都', '西安', '重庆', '天津']  # ID003, ID004的F列不同
        }

        df2 = pd.DataFrame(data2)
        df2.to_excel(self.file2_path, index=False)

        # 添加删除线格式到某些行（模拟删除的数据）
        self._add_strikethrough_format()

        print(f"已创建样例文件：{self.file1_path} 和 {self.file2_path}")

    def _add_strikethrough_format(self):
        """为表格2的某些行添加删除线格式"""
        try:
            wb = openpyxl.load_workbook(self.file2_path)
            ws = wb.active

            # 为第6行（ID006）添加删除线（整行删除）
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=6, column=col)
                cell.font = Font(strike=True)

            # 为第7行（ID007）添加删除线（整行删除）
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=7, column=col)
                cell.font = Font(strike=True)

            # 为第8行（ID008）仅部分列添加删除线（演示新的检测逻辑）
            # 只在A列和C列添加删除线
            ws.cell(row=8, column=1).font = Font(strike=True)  # A列
            ws.cell(row=8, column=3).font = Font(strike=True)  # C列

            wb.save(self.file2_path)
            print("已为样例文件添加删除线格式")
        except Exception as e:
            print(f"添加删除线格式时出错：{e}")

    def _add_strikethrough_format_table1(self):
        """为表格1的某些行添加删除线格式（用于测试）"""
        try:
            wb = openpyxl.load_workbook(self.file1_path)
            ws = wb.active

            # 为第6行（ID005）整行添加删除线（第1行是标题行）
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=6, column=col)
                cell.font = Font(strike=True)

            wb.save(self.file1_path)
            print("已为表1样例文件添加删除线格式")
        except Exception as e:
            print(f"为表1添加删除线格式时出错：{e}")

    def read_excel_with_strikethrough_check(self, file_path):
        """读取Excel文件，并检测删除线格式"""
        print(f"正在读取文件：{file_path}")

        # 使用openpyxl读取以检测格式
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # 获取表头
        headers = [cell.value for cell in ws[1]]

        # 读取数据并检查删除线格式
        data_rows = []
        strikethrough_rows = set()

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            # 检查A-Z列中是否有任意一列有删除线格式
            has_strikethrough = False

            # 只检查A-Z列（1-26列）
            max_col_to_check = min(26, len(row))  # 限制到Z列，避免越界
            for col_idx in range(max_col_to_check):
                cell = row[col_idx]
                if cell.value is not None and cell.font and cell.font.strike:
                    has_strikethrough = True
                    break

            if has_strikethrough:
                strikethrough_rows.add(row_idx - 2)  # 转换为0-based索引
                print(f"检测到删除线行：第{row_idx}行（A-Z列中有删除线）")
            else:
                # 如果没有删除线，则添加到数据中
                row_data = [cell.value for cell in row]
                data_rows.append(row_data)

        # 创建DataFrame
        df = pd.DataFrame(data_rows, columns=headers)

        return df, strikethrough_rows

    def find_differences(self):
        """比较两个Excel文件并找出差异"""
        print("\n开始比较Excel文件...")

        # 读取文件1并检测删除线
        df1, strikethrough_rows1 = self.read_excel_with_strikethrough_check(self.file1_path)

        # 读取文件2并检测删除线
        df2, strikethrough_rows2 = self.read_excel_with_strikethrough_check(self.file2_path)

        print(f"表1包含 {len(df1)} 行数据（已忽略 {len(strikethrough_rows1)} 行删除线数据）")
        print(f"表2包含 {len(df2)} 行数据（已忽略 {len(strikethrough_rows2)} 行删除线数据）")

        # 获取唯一标识列
        unique_col_name = None
        for col in df1.columns:
            if str(col).strip().startswith(self.unique_column):
                unique_col_name = col
                break

        if unique_col_name is None:
            # 如果没有找到，尝试直接使用列字母
            unique_col_name = self.unique_column
            if unique_col_name not in df1.columns:
                raise ValueError(f"无法找到唯一标识列：{self.unique_column}")

        print(f"使用 '{unique_col_name}' 列作为唯一标识")

        # 创建数据字典，以唯一标识为键
        dict1 = {}
        dict2 = {}

        for _, row in df1.iterrows():
            key = str(row[unique_col_name])
            dict1[key] = row.to_dict()

        for _, row in df2.iterrows():
            key = str(row[unique_col_name])
            dict2[key] = row.to_dict()

        # 获取所有唯一的键
        all_keys = set(dict1.keys()).union(set(dict2.keys()))

        # 找出差异
        only_in_table1 = []
        only_in_table2 = []

        print(f"\n开始比较 {self.compare_columns} 列的差异...")

        for key in all_keys:
            in_table1 = key in dict1
            in_table2 = key in dict2

            if in_table1 and not in_table2:
                only_in_table1.append(dict1[key])
                print(f"发现仅存在于表1的行：{key}")
            elif in_table2 and not in_table1:
                only_in_table2.append(dict2[key])
                print(f"发现仅存在于表2的行：{key}")
            elif in_table1 and in_table2:
                # 检查比较列是否有差异
                row1 = dict1[key]
                row2 = dict2[key]

                has_difference = False
                diff_details = []

                for col_letter in self.compare_columns:
                    # 找到对应的列名
                    col_name = None
                    for col in df1.columns:
                        if str(col).strip().startswith(col_letter):
                            col_name = col
                            break

                    if col_name and col_name in row1 and col_name in row2:
                        val1 = str(row1[col_name]) if row1[col_name] is not None else ""
                        val2 = str(row2[col_name]) if row2[col_name] is not None else ""

                        if val1 != val2:
                            has_difference = True
                            diff_details.append(f"{col_name}: '{val1}' vs '{val2}'")

                if has_difference:
                    only_in_table1.append(row1)
                    only_in_table2.append(row2)
                    print(f"发现差异行 {key}：{', '.join(diff_details)}")

        return only_in_table1, only_in_table2

    def save_results(self, only_in_table1, only_in_table2):
        """保存比较结果到Excel文件"""
        print(f"\n正在保存比较结果...")

        if only_in_table1:
            df_result1 = pd.DataFrame(only_in_table1)
            df_result1.to_excel(self.output_only_in_table1, index=False)
            print(f"已保存 {len(only_in_table1)} 行表1独有的数据到：{self.output_only_in_table1}")
        else:
            print("没有发现表1独有的数据")

        if only_in_table2:
            df_result2 = pd.DataFrame(only_in_table2)
            df_result2.to_excel(self.output_only_in_table2, index=False)
            print(f"已保存 {len(only_in_table2)} 行表2独有的数据到：{self.output_only_in_table2}")
        else:
            print("没有发现表2独有的数据")

    def run_comparison(self):
        """运行完整的比较流程"""
        print("=" * 60)
        print("Excel表格比较工具")
        print("=" * 60)

        # 检查文件是否存在，如果不存在则创建样例文件
        if not os.path.exists(self.file1_path) or not os.path.exists(self.file2_path):
            print("检测到样例文件不存在，正在创建...")
            self.create_sample_files()

        print(f"\n配置参数：")
        print(f"- 唯一标识列：{self.unique_column}")
        print(f"- 比较列：{', '.join(self.compare_columns)}")
        print(f"- 文件1：{self.file1_path}")
        print(f"- 文件2：{self.file2_path}")
        print(f"- 输出文件1：{self.output_only_in_table1}")
        print(f"- 输出文件2：{self.output_only_in_table2}")
        print()

        try:
            # 执行比较
            only_in_table1, only_in_table2 = self.find_differences()

            # 保存结果
            self.save_results(only_in_table1, only_in_table2)

            print("\n" + "=" * 60)
            print("比较完成！")
            print("=" * 60)

            # 显示摘要
            total_diff = len(only_in_table1) + len(only_in_table2)
            if total_diff > 0:
                print(f"\n摘要：")
                print(f"- 发现 {total_diff} 行差异")
                print(f"- 表1独有的行：{len(only_in_table1)} 行")
                print(f"- 表2独有的行：{len(only_in_table2)} 行")
            else:
                print("\n未发现任何差异，两个表格在指定列中内容一致。")

        except Exception as e:
            print(f"错误：{str(e)}")
            import traceback
            traceback.print_exc()


if __name__ == "__main__":
    # 检查必要的库是否已安装
    try:
        import pandas
        import openpyxl
    except ImportError as e:
        print(f"缺少必要的库：{e}")
        print("请运行以下命令安装依赖库：")
        print("pip install pandas openpyxl")
        exit(1)

    tool = ExcelCompareTool()
    tool.run_comparison()
