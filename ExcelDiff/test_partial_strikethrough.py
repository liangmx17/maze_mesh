import pandas as pd
import openpyxl
from openpyxl.styles import Font
import sys
import io

# 设置编码
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# 创建一个测试文件来验证部分删除线检测
print("创建测试文件验证部分删除线检测功能...")

# 创建测试数据
data = {
    'A': ['ID001', 'ID002', 'ID003', 'ID004'],
    'B': ['正常行', '全删除行', '部分删除行', '另一正常行'],
    'C': ['数据1', '数据2', '数据3', '数据4'],
    'D': ['描述1', '描述2', '描述3', '描述4']
}

df = pd.DataFrame(data)
df.to_excel('test_strikethrough.xlsx', index=False)

# 添加不同的删除线格式
wb = openpyxl.load_workbook('test_strikethrough.xlsx')
ws = wb.active

# 第2行：全行删除线
for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=2, column=col)
    cell.font = Font(strike=True)

# 第3行：只有B列删除线
ws.cell(row=3, column=2).font = Font(strike=True)  # B列

wb.save('test_strikethrough.xlsx')
print("测试文件已创建：test_strikethrough.xlsx")

# 测试检测功能
from excel_compare_tool import ExcelCompareTool

tool = ExcelCompareTool()
df_result, strikethrough_rows = tool.read_excel_with_strikethrough_check('test_strikethrough.xlsx')

print(f"\n检测结果：")
print(f"- 原始行数：4")
print(f"- 有效行数：{len(df_result)}")
print(f"- 删除线行数：{len(strikethrough_rows)}")
print(f"- 删除线行索引：{strikethrough_rows}")

print(f"\n有效数据：")
print(df_result.to_string())